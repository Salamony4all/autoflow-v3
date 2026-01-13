import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import shutil
import zipfile
import re
import logging

logger = logging.getLogger(__name__)

class DownloadManager:
    """Manage downloads of all generated artifacts"""
    
    def __init__(self):
        self.supported_formats = ['pdf', 'excel', 'xlsx', 'xls', 'pptx', 'zip']
    
    def get_logo_path(self):
        """Return the best available logo path"""
        candidates = [
            os.path.join('static', 'images', 'AlShaya-Logo-color@2x.png'),
            os.path.join('static', 'images', 'LOGO.png'),
            os.path.join('static', 'images', 'al-shaya-logo-white@2x.png')
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        return None
    
    def prepare_download(self, file_id, file_type, format_type, session):
        """
        Prepare file for download based on type and format
        
        Args:
            file_id: ID of the file
            file_type: Type of file (extraction, offer, presentation, mas, ve)
            format_type: Output format (pdf, excel, xlsx, xls, zip)
            session: Flask session object
        
        Returns:
            Path to the file to download
        """
        # Normalize format type
        if format_type == 'excel':
            format_type = 'xlsx'
        
        if format_type not in self.supported_formats:
            raise Exception(f'Unsupported format: {format_type}')
        
        # Get file info
        uploaded_files = session.get('uploaded_files', [])
        file_info = None
        
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                break
        
        if not file_info:
            raise Exception('File not found')
        
        session_id = session['session_id']
        
        # Handle different file types
        if file_type == 'extraction':
            return self.prepare_extraction_download(file_info, format_type, session_id)
        elif file_type == 'offer':
            return self.prepare_offer_download(file_info, format_type, session_id)
        elif file_type == 'presentation':
            return self.prepare_presentation_download(file_info, format_type, session_id)
        elif file_type == 'mas':
            return self.prepare_mas_download(file_info, format_type, session_id)
        elif file_type == 've':
            return self.prepare_ve_download(file_info, format_type, session_id)
        elif file_type == 'all':
            return self.prepare_all_downloads(file_info, session_id)
        else:
            raise Exception(f'Unknown file type: {file_type}')
    
    def prepare_extraction_download(self, file_info, format_type, session_id):
        """Prepare extracted table data for download"""
        if 'extraction_result' not in file_info:
            raise Exception('No extraction data available')
        
        extraction_result = file_info['extraction_result']
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            return self.create_extraction_excel(extraction_result, output_dir, file_info['id'])
        elif format_type == 'pdf':
            # Return existing extraction output if available
            if 'output_dir' in file_info:
                # Find PDF in output directory
                pdf_files = [f for f in os.listdir(file_info['output_dir']) if f.endswith('.pdf')]
                if pdf_files:
                    return os.path.join(file_info['output_dir'], pdf_files[0])
            raise Exception('PDF extraction not available. Generate offer first.')
        
        raise Exception(f'Format {format_type} not supported for extraction')
    
    def prepare_offer_download(self, file_info, format_type, session_id):
        """Prepare offer for download (handles both costed_data and stitched_table)"""
        # Handle multi-budget tables with stitched_table
        if 'stitched_table' in file_info and file_info.get('multibudget'):
            # Create costed_data structure from stitched_table for multi-budget
            from bs4 import BeautifulSoup
            html = file_info['stitched_table']['html']
            soup = BeautifulSoup(html, 'html.parser')
            table = soup.find('table')
            
            if not table:
                raise Exception('No table found in stitched data')
            
            # Parse table to costed_data format (excluding Product Selection and Actions columns)
            headers = []
            header_row = table.find('tr')
            if header_row:
                for th in header_row.find_all(['th', 'td']):
                    header_text = th.get_text(strip=True)
                    # Exclude Product Selection and Actions columns
                    if header_text.lower() not in ['action', 'actions', 'product selection', 'productselection']:
                        headers.append(header_text)
            
            rows = []
            seen_rows = set()  # Track unique rows for deduplication
            
            for row in table.find_all('tr')[1:]:
                cells = row.find_all('td')
                if len(cells) == 0:
                    continue
                
                row_data = {}
                col_idx = 0
                for i, cell in enumerate(cells):
                    # Skip Product Selection and Actions cells
                    if cell.find(class_='product-selection-dropdowns') or cell.find('button'):
                        continue
                    text = cell.get_text(strip=True).lower()
                    if 'product selection' in text or 'actions' in text:
                        continue
                    
                    if col_idx < len(headers):
                        # Keep image HTML if present
                        img = cell.find('img')
                        if img:
                            row_data[headers[col_idx]] = str(cell)
                        else:
                            row_data[headers[col_idx]] = cell.get_text(strip=True)
                        col_idx += 1
                
                if row_data:
                    # Create a unique ID for this row based on key columns
                    # Use description, brand, model, unit rate for uniqueness
                    key_values = []
                    for h in headers:
                        h_lower = h.lower()
                        # Only use key identifying columns for deduplication
                        if any(k in h_lower for k in ['description', 'brand', 'model', 'item', 'product', 'rate', 'price']):
                            val = row_data.get(h, '')
                            # Strip HTML from images for comparison
                            val = re.sub(r'<[^>]+>', '', str(val)).strip()
                            key_values.append(val)
                    
                    row_key = '|'.join(key_values)
                    
                    if row_key and row_key not in seen_rows:
                        rows.append(row_data)
                        seen_rows.add(row_key)
                        logger.debug(f"Added unique row with key: {row_key[:100]}...")
                    elif row_key in seen_rows:
                        logger.info(f"Skipped duplicate row: {row_key[:100]}...")
                    else:
                        # No key could be generated, add anyway
                        rows.append(row_data)
            
            logger.info(f"Excel export: {len(rows)} unique rows (removed duplicates)")
            
            # Create costed_data structure
            costed_data = {
                'tables': [{
                    'headers': headers,
                    'rows': rows
                }],
                'factors': {},
                'session_id': session_id
            }
        elif 'costed_data' not in file_info:
            raise Exception('No costed data available. Apply costing first.')
        else:
            costed_data = file_info['costed_data']
        
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            product_selections = file_info.get('product_selections', [])
            return self.create_offer_excel(costed_data, output_dir, file_info['id'], product_selections=product_selections)
        elif format_type == 'pdf':
            # Check if offer PDF was generated
            offer_dir = os.path.join('outputs', session_id, 'offers')
            if os.path.exists(offer_dir):
                pdf_files = [f for f in os.listdir(offer_dir) if f.endswith('.pdf') and file_info['id'] in f]
                if pdf_files:
                    return os.path.join(offer_dir, pdf_files[0])
            raise Exception('Offer PDF not generated yet')
        
        raise Exception(f'Format {format_type} not supported for offers')
    
    def prepare_presentation_download(self, file_info, format_type, session_id):
        """Prepare presentation for download"""
        presentation_dir = os.path.join('outputs', session_id, 'presentations')
        
        if not os.path.exists(presentation_dir):
            raise Exception('Presentation not generated yet')
        
        # Look for files with the correct extension
        if format_type == 'pptx':
            files = [f for f in os.listdir(presentation_dir) if f.endswith('.pptx') and file_info['id'] in f]
        else:  # pdf
            files = [f for f in os.listdir(presentation_dir) if f.endswith('.pdf') and file_info['id'] in f]
        
        if files:
            # Return the most recent file
            files.sort(reverse=True)
            return os.path.join(presentation_dir, files[0])
        
        raise Exception(f'Presentation file ({format_type}) not found')
    
    def prepare_mas_download(self, file_info, format_type, session_id):
        """Prepare MAS for download"""
        mas_dir = os.path.join('outputs', session_id, 'mas')
        
        if not os.path.exists(mas_dir):
            raise Exception('MAS not generated yet')
        
        pdf_files = [f for f in os.listdir(mas_dir) if f.endswith('.pdf') and file_info['id'] in f]
        if pdf_files:
            return os.path.join(mas_dir, pdf_files[0])
        
        raise Exception('MAS file not found')
    
    def prepare_ve_download(self, file_info, format_type, session_id):
        """Prepare value engineering alternatives for download"""
        if 'value_engineering' not in file_info:
            raise Exception('Value engineering not performed yet')
        
        ve_data = file_info['value_engineering']
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            return self.create_ve_excel(ve_data, output_dir, file_info['id'])
        
        raise Exception(f'Format {format_type} not supported for value engineering')
    
    def prepare_all_downloads(self, file_info, session_id):
        """Create a ZIP file with all generated documents"""
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        zip_filename = os.path.join(output_dir, f'all_documents_{file_info["id"]}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
        
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add extraction data
            if 'extraction_result' in file_info:
                try:
                    excel_file = self.create_extraction_excel(
                        file_info['extraction_result'], 
                        output_dir, 
                        file_info['id']
                    )
                    zipf.write(excel_file, os.path.basename(excel_file))
                except:
                    pass
            
            # Add offer
            if 'costed_data' in file_info:
                try:
                    offer_file = self.create_offer_excel(
                        file_info['costed_data'], 
                        output_dir, 
                        file_info['id']
                    )
                    zipf.write(offer_file, os.path.basename(offer_file))
                except:
                    pass
            
            # Add PDFs from various directories
            for subdir in ['offers', 'presentations', 'mas']:
                dir_path = os.path.join('outputs', session_id, subdir)
                if os.path.exists(dir_path):
                    for filename in os.listdir(dir_path):
                        if file_info['id'] in filename:
                            file_path = os.path.join(dir_path, filename)
                            zipf.write(file_path, f'{subdir}/{filename}')
        
        return zip_filename
    
    def create_extraction_excel(self, extraction_result, output_dir, file_id):
        """Create Excel file from extraction result"""
        filename = os.path.join(output_dir, f'extraction_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        
        # Logo removed from Excel exports per user request
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each page
        for idx, layout_result in enumerate(extraction_result.get('layoutParsingResults', [])):
            markdown_text = layout_result.get('markdown', {}).get('text', '')
            
            # Extract tables
            tables = self.parse_markdown_tables(markdown_text)
            
            for table_idx, table in enumerate(tables):
                sheet_name = f'Page{idx+1}_Table{table_idx+1}'[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                if table['headers']:
                    ws.append(table['headers'])
                    self.style_header_row(ws, 1)
                
                # Add data rows
                for row in table['rows']:
                    row_data = [row.get(h, '') for h in table['headers']]
                    ws.append(row_data)
                
                # Auto-adjust column widths
                self.auto_adjust_columns(ws)
        
        wb.save(filename)
        return filename
    
    def create_offer_excel(self, costed_data, output_dir, file_id, product_selections=None):
        """Create Excel file for offer with costing applied"""
        filename = os.path.join(output_dir, f'offer_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Offer'
        
        temp_files = []
        
        # Title
        ws.append(['COMMERCIAL OFFER'])
        ws.merge_cells('A1:F1')
        self.style_title_row(ws, 1)
        
        ws.append([])  # Empty row
        
        # Costing Factors removed - confidential information
        ws.append([])  # Empty row for spacing
        
        # Get session info for image paths
        session_id = costed_data.get('session_id', '')
        
        # Tables
        for table_idx, table in enumerate(costed_data['tables']):
            ws.append([f'Item List {table_idx + 1}'])
            ws.merge_cells(f'A{ws.max_row}:F{ws.max_row}')
            
            # Filter out Action and Product Selection columns from headers
            headers = [h for h in table['headers'] if h.lower() not in ['action', 'actions', 'product selection', 'productselection']]
            
            # Headers
            header_row_num = ws.max_row + 1
            ws.append(headers)
            self.style_header_row(ws, header_row_num)
            
            # Recalculate totals for all rows before exporting
            from utils.costing_engine import CostingEngine
            engine = CostingEngine()
            for row in table['rows']:
                row = engine.recalculate_totals(row, headers)
            
            # Find image column index
            image_col_indices = []
            for idx, h in enumerate(headers):
                if 'image' in h.lower() or 'img' in h.lower() or 'ref' in h.lower():
                    image_col_indices.append(idx)
            
            # Deduplicate rows before exporting to Excel
            unique_rows = []
            seen_rows = set()
            for row in table['rows']:
                # Create a unique key based on key columns
                key_values = []
                for h in headers:
                    h_lower = h.lower()
                    if any(k in h_lower for k in ['description', 'brand', 'model', 'item', 'product', 'rate', 'price']):
                        val = row.get(h, '')
                        val = re.sub(r'<[^>]+>', '', str(val)).strip()
                        key_values.append(val)
                
                row_key = '|'.join(key_values)
                
                if row_key and row_key not in seen_rows:
                    unique_rows.append(row)
                    seen_rows.add(row_key)
                elif not row_key:
                    unique_rows.append(row)  # No key, add anyway
            
            if len(unique_rows) < len(table['rows']):
                logger.info(f"Excel generation: Removed {len(table['rows']) - len(unique_rows)} duplicate rows")
            
            # Data rows - exclude Action column and embed images
            for row_idx, row in enumerate(unique_rows):
                current_row_num = ws.max_row + 1
                row_data = []
                
                for col_idx, h in enumerate(headers):
                    cell_value = row.get(h, '')
                    
                    # Check if this cell contains an image
                    if self.contains_image(cell_value):
                        row_data.append('')  # Empty cell, image will be placed on top
                    else:
                        # Strip HTML tags if any
                        clean_value = re.sub(r'<[^>]+>', '', str(cell_value))
                        row_data.append(clean_value)
                
                ws.append(row_data)
                
                # Set row height for images - taller if both logo and product image present
                has_image_content = any(any(v in h.lower() for v in ['image', 'img', 'picture', 'pic']) or self.contains_image(row.get(h, '')) for h in headers)
                if has_image_content:
                    if row.get('brand_logo'):
                        ws.row_dimensions[current_row_num].height = 135
                    else:
                        ws.row_dimensions[current_row_num].height = 100
                
                # Identify key columns for this row
                total_col_idx = -1
                for i, h in enumerate(headers):
                    h_lower = h.lower()
                    if any(v in h_lower for v in ['total', 'amount']) and '_original' not in h_lower:
                        total_col_idx = i + 1
                        break
                if total_col_idx == -1:
                    total_col_idx = 7 # Fallback

                # Embed images and brand logos
                for col_idx, h in enumerate(headers):
                    cell_value = str(row.get(h, ''))
                    is_image_col = any(v in h.lower() for v in ['image', 'img', 'picture', 'pic'])
                    
                    if is_image_col or self.contains_image(cell_value):
                        # Priority: get image_url from row data, then extracted path, then product_selections
                        image_url = row.get('image_url') or self.extract_image_path(cell_value, session_id, file_id)
                        brand_logo_url = row.get('brand_logo')
                        
                        # Fallback to product_selections if available (requested by user)
                        if product_selections and (not image_url or not brand_logo_url):
                            # Try to match by row index
                            selection = next((p for p in product_selections if p.get('row_index') == row_idx), None)
                            if selection:
                                if not image_url: image_url = selection.get('image_url')
                                if not brand_logo_url: brand_logo_url = selection.get('brand_logo')

                        # Final fallback for brand logo from brand name
                        if not brand_logo_url:
                            brand = row.get('brand') or row.get('Brand')
                            if brand:
                                try:
                                    from utils.image_helper import get_brand_logo_url
                                    brand_logo_url = get_brand_logo_url(brand)
                                except: pass

                        logo_added = False
                        img_added = False

                        try:
                            from utils.image_helper import download_image
                            from PIL import Image as PIL_Image
                            import tempfile
                            import traceback
                            
                            # Process Brand Logo
                            if brand_logo_url:
                                cached_logo = download_image(brand_logo_url)
                                if cached_logo and os.path.exists(cached_logo):
                                    try:
                                        # Use tempfile path to avoid 'fp' attribute error in openpyxl
                                        with PIL_Image.open(cached_logo) as pil_img:
                                            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                                            pil_img.convert('RGBA').save(tmp.name, format='PNG')
                                            tmp.close()
                                            temp_files.append(tmp.name)
                                            
                                            logo_xl = XLImage(tmp.name)
                                            logo_xl.width = 60
                                            logo_xl.height = 40
                                            
                                            marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(10), row=current_row_num-1, rowOff=pixels_to_EMU(5))
                                            # XLImage from path should have _extent, but manual set is safer
                                            ext = XDRPositiveSize2D(pixels_to_EMU(logo_xl.width), pixels_to_EMU(logo_xl.height))
                                            logo_xl.anchor = OneCellAnchor(_from=marker, ext=ext)
                                            ws.add_image(logo_xl)
                                            logo_added = True
                                            logger.info(f"Added brand logo for row {current_row_num}")
                                    except Exception as e:
                                        logger.warning(f"Error adding logo to Excel row {current_row_num}: {e}\n{traceback.format_exc()}")

                            # Process Product Image
                            if image_url:
                                cached_img = download_image(image_url) if str(image_url).startswith('http') else image_url
                                if cached_img and os.path.exists(str(cached_img)):
                                    try:
                                        with PIL_Image.open(str(cached_img)) as pil_img:
                                            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                                            pil_img.convert('RGB').save(tmp.name, format='PNG')
                                            tmp.close()
                                            temp_files.append(tmp.name)
                                            
                                            img_xl = XLImage(tmp.name)
                                            img_xl.width = 100
                                            img_xl.height = 100
                                            
                                            row_off = 50 if logo_added else 5
                                            marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(10), row=current_row_num-1, rowOff=pixels_to_EMU(row_off))
                                            ext = XDRPositiveSize2D(pixels_to_EMU(img_xl.width), pixels_to_EMU(img_xl.height))
                                            img_xl.anchor = OneCellAnchor(_from=marker, ext=ext)
                                            ws.add_image(img_xl)
                                            img_added = True
                                            logger.info(f"Added product image for row {current_row_num}")
                                    except Exception as e:
                                        logger.warning(f"Error adding product image to Excel row {current_row_num}: {e}\n{traceback.format_exc()}")
                                        if not logo_added:
                                            ws.cell(row=current_row_num, column=col_idx + 1).value = "[Img Error]"
                                else:
                                    logger.warning(f"Product image path does not exist for row {current_row_num}: {cached_img}")
                                    if is_image_col and not logo_added:
                                        ws.cell(row=current_row_num, column=col_idx + 1).value = "[Missing]"

                            # Cleanup cell text IF it's the dedicated image column
                            if is_image_col and (logo_added or img_added):
                                ws.cell(row=current_row_num, column=col_idx + 1).value = ""

                        except Exception as e:
                            logger.error(f"Excel image processing totally failed for row {current_row_num}: {e}\n{traceback.format_exc()}")

            # Update summary row logic to use identified total column
            final_summary_col = total_col_idx
            summary_label_col = final_summary_col - 1
            if summary_label_col < 1: summary_label_col = 1
            
            # Summary - recalculate to ensure accuracy
            subtotal = self.calculate_subtotal(costed_data['tables'])
            vat = subtotal * 0.15  # 15% VAT
            grand_total = subtotal + vat
            
            # Add summary rows with proper formatting
            summary_row = ws.max_row + 2
            ws.cell(row=summary_row, column=summary_label_col).value = 'Subtotal:'
            ws.cell(row=summary_row, column=final_summary_col).value = subtotal
            
            ws.cell(row=summary_row + 1, column=summary_label_col).value = 'VAT (15%):'
            ws.cell(row=summary_row + 1, column=final_summary_col).value = vat
            
            ws.cell(row=summary_row + 2, column=summary_label_col).value = 'Grand Total:'
            ws.cell(row=summary_row + 2, column=final_summary_col).value = grand_total
        
        # Style summary rows
        self.style_summary_rows(ws, summary_row, summary_row + 2)
        
        # Final column adjustments
        # First auto-adjust everything
        self.auto_adjust_columns(ws)
        
        # Then apply specific overrides for image and description columns
        for table in costed_data['tables']:
            headers = [h for h in table['headers'] if h.lower() not in ['action', 'actions', 'product selection', 'productselection']]
            for col_idx, h in enumerate(headers):
                col_letter = chr(65 + col_idx)
                if any(v in h.lower() for v in ['image', 'img', 'picture', 'pic']):
                    ws.column_dimensions[col_letter].width = 20
                elif h.lower() in ['description', 'desc']:
                    ws.column_dimensions[col_letter].width = 65
                    # Re-apply wrapping as auto_adjust might have affected it
                    for r in range(1, ws.max_row + 1):
                        cell = ws.cell(row=r, column=col_idx + 1)
                        if cell.value and r > 1: # Skip title/header
                            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        
        wb.save(filename)
        
        # Cleanup temporary files
        for tmp_file in temp_files:
            try:
                if os.path.exists(tmp_file):
                    os.remove(tmp_file)
            except: pass
            
        return filename
    
    def create_ve_excel(self, ve_data, output_dir, file_id):
        """Create Excel file for value engineering alternatives"""
        filename = os.path.join(output_dir, f've_alternatives_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Alternatives'
        
        # Logo removed from Excel exports per user request
        
        # Title
        ws.append([f'VALUE ENGINEERED ALTERNATIVES - {ve_data["budget_option"].upper()}'])
        ws.merge_cells('A3:H3')
        self.style_title_row(ws, 3)
        ws.append([])
        
        # Process alternatives
        for alt_group in ve_data['alternatives']:
            original = alt_group['original_item']
            
            # Original item
            ws.append(['ORIGINAL ITEM'])
            ws.append(['Description:', original.get('description', '')])
            ws.append(['Quantity:', f"{original.get('qty', '')} {original.get('unit', '')}"])
            ws.append(['Unit Rate:', original.get('unit_rate', '')])
            ws.append(['Total:', original.get('total', '')])
            ws.append([])
            
            # Alternatives
            ws.append(['ALTERNATIVES'])
            headers = ['Brand', 'Model', 'Description', 'Unit Rate', 'Total', 'Lead Time']
            ws.append(headers)
            self.style_header_row(ws, ws.max_row)
            
            for alt in alt_group['alternatives']:
                ws.append([
                    alt['brand'],
                    alt['model'],
                    alt['description'],
                    alt['unit_rate'],
                    alt['total'],
                    alt['lead_time']
                ])
            
            ws.append([])
            ws.append([])  # Double space between items
        
        self.auto_adjust_columns(ws)
        
        wb.save(filename)
        return filename
    
    def contains_image(self, cell_value):
        """Check if cell contains an image reference"""
        return '<img' in str(cell_value).lower() or 'img_in_' in str(cell_value).lower()
    
    def extract_image_path(self, cell_value, session_id, file_id):
        """Extract image path or URL from cell value"""
        try:
            # Look for img src pattern
            match = re.search(r'src=["\']([^"\']+)["\']', str(cell_value))
            if match:
                img_path_or_url = match.group(1)
                
                # Handle URLs
                if img_path_or_url.startswith('http'):
                    return img_path_or_url
                
                # Try to find it relative to current directory (for static/ paths)
                # Remove leading slash if any
                clean_path = img_path_or_url.lstrip('/')
                if os.path.exists(clean_path):
                    return clean_path
                
                # Check for other common paths
                if clean_path.startswith('static/'):
                    return os.path.abspath(clean_path)
                
                # Default to outputs/ (legacy behavior)
                if not clean_path.startswith('outputs'):
                    img_path = os.path.join('outputs', session_id, file_id, clean_path)
                else:
                    img_path = clean_path
                
                return img_path
            
            # Try to find image reference in text
            if 'img_in_' in str(cell_value):
                match = re.search(r'(imgs/img_in_[^"\s<>]+\.jpg)', str(cell_value))
                if match:
                    img_relative_path = match.group(1)
                    img_path = os.path.join('outputs', session_id, file_id, img_relative_path)
                    return img_path
        except Exception as e:
            logger.warning(f"Error extracting image path: {e}")
        
        return None
    
    def parse_markdown_tables(self, markdown_text):
        """Parse tables from markdown text"""
        lines = markdown_text.split('\n')
        tables = []
        current_table = {'headers': [], 'rows': []}
        in_table = False
        
        for line in lines:
            if '|' in line:
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                if not in_table:
                    # Start of table - headers
                    current_table['headers'] = cells
                    in_table = True
                elif all(c in '-|: ' for c in line):
                    # Separator line - skip
                    continue
                else:
                    # Data row
                    if len(cells) == len(current_table['headers']):
                        row = dict(zip(current_table['headers'], cells))
                        current_table['rows'].append(row)
            else:
                if in_table and current_table['rows']:
                    tables.append(current_table)
                    current_table = {'headers': [], 'rows': []}
                in_table = False
        
        if in_table and current_table['rows']:
            tables.append(current_table)
        
        return tables
    
    def calculate_subtotal(self, tables):
        """Calculate subtotal from tables"""
        import re
        subtotal = 0.0
        
        for table in tables:
            for row in table['rows']:
                for key, value in row.items():
                    if 'total' in key.lower() and '_original' not in key:
                        try:
                            cleaned = re.sub(r'[^\d.-]', '', str(value))
                            num_value = float(cleaned)
                            subtotal += num_value
                        except:
                            pass
        
        return subtotal
    
    def style_header_row(self, ws, row_num):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_title_row(self, ws, row_num):
        """Apply styling to title row"""
        title_fill = PatternFill(start_color='764BA2', end_color='764BA2', fill_type='solid')
        title_font = Font(bold=True, size=16, color='FFFFFF')
        
        for cell in ws[row_num]:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_summary_rows(self, ws, start_row, end_row):
        """Apply styling to summary rows"""
        bold_font = Font(bold=True)
        
        for row_num in range(start_row, end_row + 1):
            for cell in ws[row_num]:
                if cell.value:
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='right')
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        from openpyxl.cell.cell import MergedCell
        
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                try:
                    # Skip merged cells
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if column_letter is None:
                        column_letter = cell.column_letter
                    
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
